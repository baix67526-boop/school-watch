# watcher.py  (带 DEBUG 输出版：用于确认 ALWAYS_SEND_SUMMARY / TEST_MAIL_TO 是否生效)
# 功能：
# 1) 读取 sources.txt（每行：学校名 URL）
# 2) 定时抓取栏目页，做“降噪指纹”对比，判断是否有更新
# 3) 读取 subscriptions.xlsx（email/schools/status），只给订阅了该学校的人发邮件（逐个单发）
# 4) 汇总验收开关：
#    - ALWAYS_SEND_SUMMARY=1 且设置 TEST_MAIL_TO 时：每次运行都给 TEST_MAIL_TO 发一封【监控汇总】（即使无更新）
# 5) 将最新指纹写回 state.json（配合 Actions 提交回仓库）
#
# DEBUG：
# - 在 send_summary_email() 会打印：
#   DEBUG ALWAYS_SEND_SUMMARY = '...'
#   DEBUG TEST_MAIL_TO set? = YES/NO
#   DEBUG sending summary to TEST_MAIL_TO...

import os
import json
import time
import hashlib
import smtplib
from email.mime.text import MIMEText
from email.header import Header
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import load_workbook

STATE_FILE = "state.json"
SOURCES_FILE = "sources.txt"
SUBS_FILE = "subscriptions.xlsx"


# ------------------- HTTP Session（带重试） -------------------
def make_session():
    s = requests.Session()
    retry = Retry(
        total=3,
        backoff_factor=1.2,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"],
        raise_on_status=False,
    )
    s.mount("http://", HTTPAdapter(max_retries=retry))
    s.mount("https://", HTTPAdapter(max_retries=retry))
    return s


SESSION = make_session()


# ------------------- 邮件发送（126：SMTP SSL 465） -------------------
def send_email_to(subject: str, body: str, to_email: str):
    smtp_host = os.getenv("SMTP_HOST", "smtp.126.com")
    smtp_port = int(os.getenv("SMTP_PORT", "465"))
    smtp_user = os.getenv("SMTP_USER")   # 你的126邮箱地址
    smtp_pass = os.getenv("SMTP_PASS")   # 你的126邮箱授权码（只放Secrets）
    if not smtp_user or not smtp_pass:
        raise RuntimeError("Missing SMTP_USER / SMTP_PASS in env")

    msg = MIMEText(body, "plain", "utf-8")
    msg["From"] = Header(smtp_user, "utf-8")
    msg["To"] = Header(to_email, "utf-8")
    msg["Subject"] = Header(subject, "utf-8")

    with smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=30) as server:
        server.login(smtp_user, smtp_pass)
        server.sendmail(smtp_user, [to_email], msg.as_string())


# ------------------- 验收汇总邮件开关（带 DEBUG） -------------------
def send_summary_email(updates_by_school, failures):
    """
    每次运行发送一封“监控汇总”到 TEST_MAIL_TO（验收用）。
    只有 ALWAYS_SEND_SUMMARY=1 且设置了 TEST_MAIL_TO 才会发送。
    """
    always = str(os.getenv("ALWAYS_SEND_SUMMARY", "")).strip()
    test_to = (os.getenv("TEST_MAIL_TO") or "").strip()

    # DEBUG（不打印具体邮箱内容，避免泄露）
    print("DEBUG ALWAYS_SEND_SUMMARY =", repr(always))
    print("DEBUG TEST_MAIL_TO set? =", "YES" if test_to else "NO")

    if always != "1":
        return
    if not test_to:
        # 没设置测试收件人就不发，避免误发
        return

    print("DEBUG sending summary to TEST_MAIL_TO...")

    now_str = time.strftime("%Y-%m-%d %H:%M", time.localtime())
    updated_schools = list(updates_by_school.keys())
    ok_count = len(updated_schools)
    fail_count = len(failures)

    lines = []
    lines.append(f"本次监控汇总（{now_str}）")
    lines.append("")
    lines.append(f"- 有更新学校数：{ok_count}")
    lines.append(f"- 抓取失败数：{fail_count}")
    lines.append("")

    if updated_schools:
        lines.append("【更新学校】")
        for school in updated_schools:
            lines.append(f"- {school}")
        lines.append("")

    if failures:
        lines.append("【失败列表（节选）】")
        for i, (name, url, err) in enumerate(failures[:10], start=1):
            lines.append(f"{i}. {name}  {url}")
            lines.append(f"   {err}")
        if len(failures) > 10:
            lines.append(f"... 还有 {len(failures)-10} 条未展示")
        lines.append("")

    lines.append("提示：这是验收/健康度汇总邮件，用于确认定时任务与发信链路正常。")
    body = "\n".join(lines)
    subject = f"【监控汇总】更新{ok_count}｜失败{fail_count}｜{now_str}"
    send_email_to(subject, body, test_to)


# ------------------- 抓取 + 降噪指纹 -------------------
def fetch(url: str) -> str:
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) SchoolWatcher/1.0",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
    }
    r = SESSION.get(url, headers=headers, timeout=25)
    r.raise_for_status()
    r.encoding = r.apparent_encoding or r.encoding
    return r.text


def normalize_html(html: str) -> str:
    """
    降噪：去 script/style，提取纯文本；只取前400行减少页脚/访问量等变化导致误报
    """
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    text = soup.get_text(separator="\n")
    lines = [ln.strip() for ln in text.splitlines()]
    lines = [ln for ln in lines if ln]
    return "\n".join(lines[:400])


def fingerprint(content: str) -> str:
    return hashlib.sha256(content.encode("utf-8")).hexdigest()


# ------------------- 文件读写 -------------------
def load_state():
    if not os.path.exists(STATE_FILE):
        return {}
    with open(STATE_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_state(state):
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


def load_sources():
    """
    sources.txt 每行格式：学校名 URL
    例如：北京大学 https://xxx...
    """
    items = []
    with open(SOURCES_FILE, "r", encoding="utf-8") as f:
        for ln in f:
            ln = ln.strip()
            if not ln or ln.startswith("#"):
                continue
            parts = ln.split()
            if len(parts) < 2:
                continue
            name = parts[0].strip()
            url = parts[1].strip()
            items.append({"name": name, "url": url})
    return items


# ------------------- 读取订阅Excel -------------------
def load_subscriptions():
    """
    subscriptions.xlsx 表头要求：
    email | schools | status

    schools：用逗号分隔学校名，例如：北京大学,清华大学
    status：ACTIVE 才发
    """
    if not os.path.exists(SUBS_FILE):
        raise RuntimeError(f"Missing {SUBS_FILE} in repo root")

    wb = load_workbook(SUBS_FILE)
    ws = wb.active

    header = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            header[str(cell.value).strip()] = idx

    for col in ["email", "schools", "status"]:
        if col not in header:
            raise RuntimeError(f"{SUBS_FILE} 缺少表头列：{col}")

    school_to_emails = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        email = row[header["email"] - 1]
        schools = row[header["schools"] - 1]
        status = row[header["status"] - 1]

        if not email or not schools:
            continue
        if str(status).strip().upper() != "ACTIVE":
            continue

        email = str(email).strip()
        raw = str(schools).replace("，", ",")
        school_list = [s.strip() for s in raw.split(",") if s.strip()]

        for s in school_list:
            school_to_emails.setdefault(s, set()).add(email)

    return school_to_emails


# ------------------- 主流程 -------------------
def main():
    sources = load_sources()
    state = load_state()
    school_to_emails = load_subscriptions()

    updates_by_school = {}  # {school_name: [url1, url2, ...]}
    failures = []  # (school, url, error)

    for item in sources:
        name, url = item["name"], item["url"]
        try:
            html = fetch(url)
            norm = normalize_html(html)
            fp = fingerprint(norm)

            old_fp = state.get(url, {}).get("fp")
            if old_fp and old_fp != fp:
                updates_by_school.setdefault(name, []).append(url)

            state[url] = {"fp": fp, "ts": int(time.time())}
        except Exception as e:
            failures.append((name, url, str(e)))
            state[url] = {
                "fp": state.get(url, {}).get("fp"),
                "ts": int(time.time()),
                "error": str(e),
            }

    save_state(state)

    # 每次运行给 TEST_MAIL_TO 发“监控汇总”（验收用）
    send_summary_email(updates_by_school, failures)

    # 没更新就不打扰订阅用户（汇总邮件已发给你自己）
    if not updates_by_school:
        print("No updates.")
        return

    # 按收件人汇总要发的内容（逐个单发）
    email_to_school_urls = {}  # {email: {school: [urls...]}}
    for school, urls in updates_by_school.items():
        subs = school_to_emails.get(school, set())
        for em in subs:
            email_to_school_urls.setdefault(em, {}).setdefault(school, []).extend(urls)

    if not email_to_school_urls:
        print("Updates exist, but no subscribers matched.")
        return

    now_str = time.strftime("%Y-%m-%d %H:%M", time.localtime())
    for em, school_map in email_to_school_urls.items():
        lines = [f"检测到你订阅的学校栏目有更新（{now_str}）：", ""]
        for school, urls in school_map.items():
            lines.append(f"")
            for u in urls:
                host = urlparse(u).netloc
                lines.append(f"- {host}\n  {u}")
            lines.append("")
        lines.append("提示：请以学校官网为准。若需暂停/退订，请微信联系我。")

        subject = f"【调剂信息更新提醒】{len(school_map)}所学校栏目有变化"
        body = "\n".join(lines).strip()

        send_email_to(subject, body, em)
        print(f"Sent to {em}: {len(school_map)} schools updated.")


if __name__ == "__main__":
    main()
